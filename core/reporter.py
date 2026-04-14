"""
core/reporter.py
Generates two Excel files:
  1. similarity_matrix_TIMESTAMP.xlsx  — N×N colour-coded matrix
  2. flagged_pairs_TIMESTAMP.xlsx      — pairs above threshold, DESCENDING sort,
                                         with raw text snippets
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_THIN    = Border(**{s: Side(style="thin", color="CCCCCC")
                     for s in ("left","right","top","bottom")})
_NO_FILL = PatternFill(fill_type=None)
_HEADER  = PatternFill("solid", fgColor="1E293B")
_RED     = PatternFill("solid", fgColor="FCA5A5")
_ORANGE  = PatternFill("solid", fgColor="FCD34D")
_YELLOW  = PatternFill("solid", fgColor="FEF08A")
_GREY    = PatternFill("solid", fgColor="E2E8F0")


def _fill(score: float) -> PatternFill:
    if score >= 90: return _RED
    if score >= 70: return _ORANGE
    if score >= 50: return _YELLOW
    return _NO_FILL


def _hcell(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.fill      = _HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _THIN
    return cell


def _dcell(ws, r, c, val, fill=None, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _THIN
    cell.fill      = fill or _NO_FILL
    if bold:
        cell.font  = Font(bold=True, name="Calibri", size=10)
    return cell


def generate(
    output_dir: Path,
    timestamp:  str,
    rolls:      list[str],
    pair_results: list[dict],   # from comparator.compare() — already sorted descending
    raw_texts:  dict[str, str],
    threshold:  float,
) -> tuple[Path, Path]:

    output_dir.mkdir(parents=True, exist_ok=True)
    matrix_path  = output_dir / f"similarity_matrix_{timestamp}.xlsx"
    flagged_path = output_dir / f"flagged_pairs_{timestamp}.xlsx"

    # ── Score lookup ─────────────────────────────────────────────────────────
    score_map: dict[tuple, dict] = {}
    for r in pair_results:
        score_map[(r["roll_a"], r["roll_b"])] = r
        score_map[(r["roll_b"], r["roll_a"])] = r

    # ── Matrix ────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Similarity Matrix"

    _hcell(ws, 1, 1, "Roll ↓ \\ →")
    ws.column_dimensions["A"].width = 15

    for ci, roll in enumerate(rolls, start=2):
        _hcell(ws, 1, ci, roll)
        ws.column_dimensions[get_column_letter(ci)].width = 13

    for ri, roll_a in enumerate(rolls, start=2):
        _hcell(ws, ri, 1, roll_a)
        ws.row_dimensions[ri].height = 18
        for ci, roll_b in enumerate(rolls, start=2):
            if roll_a == roll_b:
                _dcell(ws, ri, ci, "—", _GREY)
            else:
                entry = score_map.get((roll_a, roll_b))
                score = entry["similarity"] if entry else 0.0
                _dcell(ws, ri, ci, f"{score}%", _fill(score),
                       bold=score >= threshold)

    ws.freeze_panes = "B2"
    wb.save(matrix_path)

    # ── Flagged pairs (descending — already sorted by comparator) ─────────────
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Flagged Pairs"

    headers = ["Rank", "Student A", "Student B", "Similarity %",
               "SequenceMatcher %", "FuzzyToken %",
               "Student A Code", "Student B Code"]
    widths  = [7, 15, 15, 14, 18, 14, 55, 55]

    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        _hcell(ws2, 1, ci, h)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    flagged = [r for r in pair_results if r["similarity"] >= threshold]
    # pair_results already descending — flagged inherits that order

    if not flagged:
        ws2.cell(row=2, column=1,
                 value="No pairs above threshold.").font = Font(italic=True)
    else:
        for rank, r in enumerate(flagged, start=1):
            fill = _fill(r["similarity"])
            ri   = rank + 1

            _dcell(ws2, ri, 1, rank, fill)
            _dcell(ws2, ri, 2, r["roll_a"], fill)
            _dcell(ws2, ri, 3, r["roll_b"], fill)
            _dcell(ws2, ri, 4, f"{r['similarity']}%",  fill, bold=True)
            _dcell(ws2, ri, 5, f"{r['seq_score']}%",   fill)
            _dcell(ws2, ri, 6, f"{r['fuzzy_score']}%", fill)

            for col, roll in [(7, r["roll_a"]), (8, r["roll_b"])]:
                snippet = (raw_texts.get(roll) or "")[:1500]
                c = ws2.cell(row=ri, column=col, value=snippet)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border    = _THIN

            ws2.row_dimensions[ri].height = 90

    ws2.freeze_panes = "A2"
    wb2.save(flagged_path)
    return matrix_path, flagged_path
